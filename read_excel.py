import openpyxl
import openpyxl.utils.exceptions
import sys


# Класс, представляющий прочитанную таблицу
class Table:
    def __init__(self, name, table):
        self.name = name
        self.table = table


# Возвращает список таблиц из файла filename (для каждого листа - своя таблица)
def get_tables(filename):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        print(f"Файл {filename} не найден")
        sys.exit(1)
    except openpyxl.utils.exceptions.InvalidFileException:
        print(f"Файл {filename} не является файлом формата xlsx")
        sys.exit(1)
    except PermissionError:
        print(f"Нет доступа к файлу {filename}")
        sys.exit(1)
    except OSError:
        print(f"Файл {filename} поврежден")
        sys.exit(1)

    tables = []
    for sheet in wb.worksheets:
        max_row = sheet.max_row
        max_column = sheet.max_column
        min_row = sheet.min_row
        min_column = sheet.min_column

        # Считываем очередную таблицу
        table = []
        for i in range(min_row, max_row + 1):
            row = []
            for j in range(min_column, max_column + 1):
                cell = sheet.cell(row=i, column=j)
                row.append(cell.value)
            table.append(row)

        has_values = False
        for cell in table[-1]:
            if cell is not None:
                has_values = True
                break
        if not has_values:
            table.pop()

        has_values = False
        for row in table:
            if row[-1] is not None:
                has_values = True
                break
        if not has_values:
            for row in table:
                row.pop()

        tables.append(Table(f"{filename[0:str(filename).rfind('.')]}.{sheet.title}", table))
    print(f"Таблицы из файла {filename} успешно прочитаны")
    return tables
